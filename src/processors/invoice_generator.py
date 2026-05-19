# Este es invoice_generator.py

import io
import zipfile
from datetime import datetime
from typing import Dict, Any, List

import streamlit as st
import pandas as pd
from openpyxl import load_workbook

# ----------------------------
# FUNCIONES AUXILIARES
# ----------------------------
def safe_float(v, default=0.0):
    """Convierte un valor a float de forma segura, devolviendo un valor por defecto en caso de error."""
    try:
        return float(v)
    except (ValueError, TypeError):
        return default

def fmt_invoice_number_from_seq(seq: int) -> str:
    """Formatea un número de secuencia de factura al formato estándar."""
    return f"002-001-01-{seq:08d}"

# ----------------------------
# CARGADOR DE PRODUCTOS
# ----------------------------
@st.cache_data
def load_products_from_template(template_path: str) -> pd.DataFrame:
    """Carga la lista de productos y servicios desde la hoja 'PRODUCTOS Y SERVICIOS' de la plantilla Excel."""
    try:
        df = pd.read_excel(template_path, sheet_name="PRODUCTOS Y SERVICIOS")
        df = df.dropna(subset=['DESCRIPCION', 'PRECIO'])
        df['PRECIO'] = df['PRECIO'].apply(lambda x: safe_float(x, 0.0))
        return df
    except Exception as e:
        st.error(f"No se pudo cargar la hoja 'PRODUCTOS Y SERVICIOS'. Error: {e}")
        return pd.DataFrame(columns=['DESCRIPCION', 'PRECIO'])

# ----------------------------
# GENERADOR DE DATOS DE FACTURA
# ----------------------------
def generate_invoice_data(validated_data: Dict[str, Any], selected_items: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Prepara el diccionario de datos para una factura individual a partir de los ítems seleccionados."""
    items = []
    for item_data in selected_items:
        is_taxable = "SERIE DUA" not in item_data.get("descripcion", "")
        cantidad = item_data.get("cantidad", 1)
        precio_unitario = item_data.get("precio_unitario", 0.0)
        
        item = {
            "cantidad": cantidad,
            "descripcion": item_data.get("descripcion"),
            "precio_unitario": precio_unitario,
            "total": cantidad * precio_unitario,
            "taxable": is_taxable,
        }
        items.append(item)

    importe_gravado = sum(item["total"] for item in items if item["taxable"])
    importe_exento = sum(item["total"] for item in items if not item["taxable"])
    sub_total = importe_gravado + importe_exento
    isv = importe_gravado * 0.15
    total = sub_total + isv

    return {
        "items": items,
        "sub_total": sub_total,
        "importe_gravado_15": importe_gravado,
        "importe_exento": importe_exento,
        "isv_15": isv,
        "total_a_pagar": total,
        "cliente": validated_data.get("cliente_preview") or validated_data.get("bl_data", {}).get("consignee_details", {}).get("name", "INVERSIONES Y SERVICIOS TÉCNICOS Y LOGISTICOS, S. DE R.L."),
        "rtn": validated_data.get("rtn_preview") or "05019021248075",
        "direccion": validated_data.get("direccion_preview") or validated_data.get("bl_data", {}).get("consignee_details", {}).get("address", "SAN PEDRO SULA, CORTES"),
        "fecha": datetime.now().strftime("%d de %B de %Y"),
        "fecha_vencimiento": (datetime.now() + pd.DateOffset(months=1)).strftime("%d de %B de %Y"),
        "factura_n": "002-001-01-00002294",
        "cai": "2BE8CD-E57AD3-2146E0-63BE03-090981-28",
        "bl_number": validated_data.get("bl_data", {}).get("bl_number", ""),
    }

# ----------------------------
# CLASE PRINCIPAL
# ----------------------------
class InvoiceGenerator:
    def __init__(self, template_path: str = "FACTURA MODELO.xlsm"):
        self.template_path = template_path
        for key, default in {
            "invoice_data_list": [], "invoice_excel_output": None, 
            "is_zip_file": False, "item_quantities": {},
            "excel_ready": False
        }.items():
            if key not in st.session_state:
                st.session_state[key] = default
        
        self.products_df = load_products_from_template(self.template_path)

    def create_invoice_excel(self, invoice_data: Dict[str, Any]) -> bytes:
        try:
            wb = load_workbook(filename=self.template_path, keep_vba=True)
        except FileNotFoundError:
            st.error(f"Plantilla no encontrada: {self.template_path}")
            return b""

        ws = wb["FACTURA"] if "FACTURA" in wb.sheetnames else wb.active
        ws_clientes = wb["CLIENTES"] if "CLIENTES" in wb.sheetnames else None
        ws_registros = wb["REGISTROS"] if "REGISTROS" in wb.sheetnames else None

        # Manejo de secuencia
        seq_new = None
        if ws_registros is not None:
            try:
                seq_current = int(ws_registros["D2"].value or 0)
            except Exception:
                seq_current = 0
            seq_new = seq_current + 1

        factura_formatted = fmt_invoice_number_from_seq(seq_new) if seq_new is not None else invoice_data.get("factura_n")
        ws["A3"] = f"CAI: {invoice_data.get('cai','')}"
        ws["C4"] = datetime.now().date()
        ws["C6"] = invoice_data.get("cliente", "")
        ws["C7"] = invoice_data.get("rtn", "")
        if invoice_data.get("direccion"):
            ws["C8"] = invoice_data.get("direccion", "")
        if factura_formatted:
            ws["A2"] = f"FACTURA     PTO CORTES      {factura_formatted}"
            ws["F2"] = factura_formatted

        # Inserta cliente si no existe
        if ws_clientes is not None:
            client_name = invoice_data.get("cliente", "").strip()
            client_rtn = invoice_data.get("rtn", "").strip()
            client_dir = invoice_data.get("direccion", "").strip()
            exists = False
            for row_idx in range(2, ws_clientes.max_row + 1):
                name_cell = ws_clientes.cell(row=row_idx, column=1).value
                rtn_cell = ws_clientes.cell(row=row_idx, column=2).value
                if name_cell and client_name and str(name_cell).strip().upper() == client_name.upper():
                    exists = True; break
                if rtn_cell and client_rtn and str(rtn_cell).strip() == client_rtn:
                    exists = True; break
            if not exists:
                ws_clientes.insert_rows(2)
                ws_clientes["A2"] = client_name; ws_clientes["B2"] = client_rtn; ws_clientes["C2"] = client_dir

        # Registros
        if ws_registros is not None and seq_new is not None:
            ws_registros.insert_rows(2)
            ws_registros["A2"] = invoice_data.get("cliente", "")
            ws_registros["B2"] = invoice_data.get("rtn", "")
            ws_registros["C2"] = invoice_data.get("bl_number", "")
            ws_registros["D2"] = seq_new
            ws_registros["E2"] = datetime.now().date()
            ws_registros["F2"] = safe_float(invoice_data.get("sub_total", 0.0))
            ws_registros["G2"] = safe_float(invoice_data.get("isv_15", 0.0))
            ws_registros["H2"] = safe_float(invoice_data.get("total_a_pagar", 0.0))

        # Limpia rango de items
        start_row, end_row = 13, 28
        for merged_range in list(ws.merged_cells.ranges):
            if merged_range.min_row <= end_row and merged_range.max_row >= start_row:
                ws.unmerge_cells(str(merged_range))
        for r in range(start_row, end_row + 1):
            for c in range(1, 8+1):
                ws.cell(row=r, column=c).value = None

        # Inserta items
        for i, it in enumerate(invoice_data.get("items", [])):
            if i >= (end_row - start_row + 1): break
            r = start_row + i
            ws[f"A{r}"] = it.get("descripcion", "")
            ws[f"D{r}"] = it.get("cantidad", 1)
            ws[f"E{r}"] = safe_float(it.get("precio_unitario", 0.0))
            ws[f"F{r}"] = safe_float(it.get("descuento", 0.0))
            ws[f"H{r}"] = f"=IF(A{r}=\"\",\"\",(D{r}*E{r})-F{r})"
            ws[f"E{r}"].number_format = '"L."#,##0.00'
            ws[f"H{r}"].number_format = '"L."#,##0.00'

        # Totales
        currency_format = '"L."#,##0.00'
        ws["H29"] = "=SUM(H13:H28)"; ws["H29"].number_format = currency_format
        ws["H31"] = '=SUMIF(A13:A28, "*SERIE DUA*", H13:H28)'; ws["H31"].number_format = currency_format
        ws["H32"] = "=H29-H31"; ws["H32"].number_format = currency_format
        ws["H34"] = "=H32*0.15"; ws["H34"].number_format = currency_format
        ws["H36"] = "=H29+H34"; ws["H36"].number_format = currency_format
        ws["F29"], ws["F31"], ws["F32"], ws["F34"], ws["F36"] = ("SUB TOTAL", "IMPORTE EXENTO", "IMPORTE GRAVADO 15%", "ISV 15%", "TOTAL A PAGAR")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # ----------------------------------------------------
    # INTERFAZ STREAMLIT CON VISTA PREVIA EDITABLE
    # ----------------------------------------------------
    def render_invoice_interface(self):
        st.subheader("Paso 3: Generar Factura de Venta")
        if not st.session_state.get("validated_data"):
            st.warning("Primero debe completar la validación en la pestaña 1.", icon="⚠️")
            return

        st.info("Defina la cantidad para cada producto que desea incluir en la factura.")

        # Selección inicial
        num_columns = 3
        cols = st.columns(num_columns)
        for i, product in enumerate(self.products_df.itertuples()):
            col = cols[i % num_columns]
            with col:
                with st.container(border=True):
                    st.markdown(f"**{product.DESCRIPCION}**")
                    st.markdown(f"L. {product.PRECIO:,.2f}")
                    st.session_state.item_quantities[product.DESCRIPCION] = st.number_input(
                        "Cantidad", min_value=0,
                        value=st.session_state.item_quantities.get(product.DESCRIPCION, 0),
                        step=1, key=f"qty_{product.DESCRIPCION}",
                        label_visibility="collapsed"
                    )

        st.markdown("---")

        if st.button("Generar Vista Previa de Factura", type="primary", use_container_width=True):
            items_to_invoice = {
                desc: qty for desc, qty in st.session_state.item_quantities.items() if qty > 0
            }

            if not items_to_invoice:
                st.error("Debe definir una cantidad (mayor a 0) para al menos un producto.")
                return

            # Filtra productos seleccionados
            selected_df = self.products_df[self.products_df['DESCRIPCION'].isin(items_to_invoice.keys())].copy()
            selected_df['cantidad'] = selected_df['DESCRIPCION'].apply(lambda x: items_to_invoice.get(x, 0))

            # --- CIF / FOB ---
            validated_data = st.session_state.get("validated_data", {})
            bl_data = validated_data.get("bl_data", {})
            all_invoices_data = validated_data.get("all_invoices_data", [])
            incoterm = ""
            if all_invoices_data and isinstance(all_invoices_data, list) and all_invoices_data[0]:
                incoterm = all_invoices_data[0].get('incoterm', '').upper()
            freight = safe_float(bl_data.get('freight_cost', 0.0))
            is_cif = 'CIF' in incoterm or (not incoterm and freight > 0)
            proration_factor = 1
            if is_cif: 
                st.success("Modo CIF detectado. Prorrateando costos de flete y seguro entre los ítems de servicio.")
                total_fob_value = (selected_df['PRECIO'] * selected_df['cantidad']).sum()
                insurance = total_fob_value * 0.015
                total_extra_cost = freight + insurance
                total_cif_value = total_fob_value + total_extra_cost
                proration_factor = total_cif_value / total_fob_value if total_fob_value > 0 else 1
                selected_df['PRECIO'] = selected_df['PRECIO'] * proration_factor
                with st.expander("Detalles del Cálculo de Prorrateo CIF"):
                    st.metric("Valor FOB Total", f"L. {total_fob_value:,.2f}")
                    st.metric("Flete del BL", f"L. {freight:,.2f}")
                    st.metric("Seguro (1.5%)", f"L. {insurance:,.2f}")
                    st.metric("Factor prorrateo aplicado", f"{proration_factor:.6f}")
            else:
                st.info("Modo FOB detectado. No se aplica prorrateo adicional.")

            # Lista de ítems seleccionados
            full_selected_items = selected_df.rename(
                columns={"DESCRIPCION": "descripcion", "PRECIO": "precio_unitario"}
            ).to_dict('records')

            # --- Vista Previa ---
            st.subheader("Vista Previa Editable de Factura")
            for i, item in enumerate(full_selected_items):
                cols = st.columns([3, 2, 1])  # Nombre | Precio | Cantidad editable
                with cols[0]:
                    st.text(item['descripcion'])
                with cols[1]:
                    st.text(f"L. {item['precio_unitario']:,.2f}")
                with cols[2]:
                    key = f"preview_qty_{i}"
                    item['cantidad'] = st.number_input(
                        "", min_value=0, value=item['cantidad'],
                        key=key, label_visibility="collapsed"
                    )

            # --- Datos Cliente ---
            st.subheader("Datos del Cliente")
            cliente = st.text_input("Nombre del Cliente", validated_data.get("bl_data", {}).get("consignee_details", {}).get("name", ""))
            rtn = st.text_input("RTN", validated_data.get("bl_data", {}).get("consignee_details", {}).get("rtn", ""))
            direccion = st.text_area("Dirección", validated_data.get("bl_data", {}).get("consignee_details", {}).get("address", ""))

            st.session_state.validated_data['cliente_preview'] = cliente
            st.session_state.validated_data['rtn_preview'] = rtn
            st.session_state.validated_data['direccion_preview'] = direccion

            # Guardar selección para Excel
            st.session_state.full_selected_items_preview = full_selected_items
            st.session_state.is_cif_preview = is_cif
            st.session_state.proration_factor_preview = proration_factor if is_cif else 1

            # --- Generación Automática del Excel ---
            invoice_data = generate_invoice_data(st.session_state.validated_data, full_selected_items)
            excel_bytes = self.create_invoice_excel(invoice_data)
            if excel_bytes:
                st.session_state["excel_ready"] = True
                st.session_state["invoice_excel_output"] = excel_bytes
                st.session_state["invoice_file_name"] = f"Factura_{invoice_data.get('factura_n','')}.xlsm"
                st.success("✅ Factura generada automáticamente. Puedes descargarla abajo.")

        # =================================================
        # BOTÓN DE DESCARGA (siempre al final)
        # =================================================
        st.markdown("---")
        if st.session_state.get("invoice_excel_output"):
            st.download_button(
                label="📥 Descargar Factura Excel",
                data=st.session_state["invoice_excel_output"],
                file_name=st.session_state.get("invoice_file_name", "Factura.xlsm"),
                mime="application/vnd.ms-excel",
                use_container_width=True
            )
        else:
            st.info("⚠️ Aún no se ha generado ninguna factura. Selecciona productos y genera la vista previa.")
